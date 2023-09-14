import math

import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1, value='время в секундах')

inp =[512,715,829,853,1014,872,886,861,852,787,857,838,810,839,808,821,833,868,847,899,843,839,835,826,827,814,816,819,816,776,781,773,789,777,782,795,797,801,765,796,862,869,825,796,809,739,743,872,811,789,801,752,784,787,820,843,892,948,923,931,895,894,852,835,861,841,862,873,2305,980,888,939,908,959,874,776,791,890,855,904,898,894,901,904,891,948,888,928,1012,837,800,843,864,854,810,784,835,844,912,810,844,892,780,797,827,799,801,843,819,780,859,841,827,806,868,863,803,779,787,853,798,813,833,654,727,823,827,854,868,846,817,849]
numbers = inp
for i in range(16):
    numbers.remove(max(numbers))
    numbers.remove(min(numbers))

print(len(numbers))
for row_idx, number in enumerate(numbers, start=2):
    sheet.cell(row=row_idx, column=1, value=number/1000)
workbook.save('output.xlsx')
workbook.close()

print("complete")
