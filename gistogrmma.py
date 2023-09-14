import math
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"

inp =[512,715,829,853,1014,872,886,861,852,787,857,838,810,839,808,821,833,868,847,899,843,839,835,826,827,814,816,819,816,776,781,773,789,777,782,795,797,801,765,796,862,869,825,796,809,739,743,872,811,789,801,752,784,787,820,843,892,948,923,931,895,894,852,835,861,841,862,873,2305,980,888,939,908,959,874,776,791,890,855,904,898,894,901,904,891,948,888,928,1012,837,800,843,864,854,810,784,835,844,912,810,844,892,780,797,827,799,801,843,819,780,859,841,827,806,868,863,803,779,787,853,798,813,833,654,727,823,827,854,868,846,817,849]
numbers = []

for i in inp:
    numbers.append(int(i))

for i in range(16):
    numbers.remove(max(numbers))
    numbers.remove(min(numbers))

max_in_inp = max(numbers)
min_in_inp = min(numbers)
delta = ((max_in_inp-min_in_inp)/math.sqrt(len(numbers)))
print((numbers),'vhdf')
print(delta, 'среднее значение')
first = []
second = []
third = []
forth = []
fifth = []
sixth = []
seventh = []
eighth = []
ninth = []
tenth = []

for i in numbers:
    if i >= min_in_inp and i < min_in_inp + delta * 1:
        first.append(i)
    elif i >= min_in_inp + delta * 1 and i < min_in_inp + delta * 2:
        second.append(i)
    elif i >= min_in_inp +delta * 2 and i < min_in_inp + delta * 3:
        third.append(i)
    elif i >= min_in_inp +delta * 3 and i < min_in_inp + delta * 4:
        forth.append(i)
    elif i >= min_in_inp +delta * 4 and i < min_in_inp + delta * 5:
        fifth.append(i)
    elif i >= min_in_inp +delta * 5 and i < min_in_inp + delta * 6:
        sixth.append(i)
    elif i >= min_in_inp +delta * 6 and i < min_in_inp + delta * 7:
        seventh.append(i)
    elif i >= min_in_inp +delta * 7 and i < min_in_inp + delta * 8:
        eighth.append(i)
    elif i >= min_in_inp +delta * 8 and i < min_in_inp + delta * 9:
        ninth.append(i)
    elif i >= min_in_inp +delta * 9 and i <= min_in_inp + delta * 10:
        tenth.append(i)

answer = first + second + third + forth + fifth +sixth + seventh + eighth +ninth + tenth
#
# for i in numbers:
#     if i not in answer:
#         print(i)

print(len(first),'1')
print(len(second),'2')
print(len(third),'3')
print(len(forth),'4')
print(len(fifth),'5')
print(len(sixth),'6')
print(len(seventh),'7')
print(len(eighth),'8')
print(len(ninth),'9')
print(len(tenth),'10')
print(len(first)+len(second)+len(third)+len(forth)+len(fifth)+len(sixth)+len(seventh)+len(eighth)+len(ninth)+len(tenth),'всего чисел')
print(first)
print(second)
print(third)
print(forth)
print(fifth)
print(sixth)
print(seventh)
print(eighth)
print(ninth)
print(tenth)

print(len(answer))
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# for row_idx, first in enumerate(first, start=2):
#     sheet.cell(row=row_idx, column=1, value=first)
#
# for row_idx, second in enumerate(second, start=2):
#     sheet.cell(row=row_idx, column=2, value=second)
#
# for row_idx, third in enumerate(third, start=2):
#     sheet.cell(row=row_idx, column=3, value=third)
#
# for row_idx, forth in enumerate(forth, start=2):
#     sheet.cell(row=row_idx, column=4, value=forth)
#
# for row_idx, fifth in enumerate(fifth, start=2):
#     sheet.cell(row=row_idx, column=5, value=fifth)
#
# for row_idx, sixth in enumerate(sixth, start=2):
#     sheet.cell(row=row_idx, column=6, value=sixth)
#
# for row_idx, seventh in enumerate(seventh, start=2):
#     sheet.cell(row=row_idx, column=7, value=seventh)
#
# for row_idx, eighth in enumerate(eighth, start=2):
#     sheet.cell(row=row_idx, column=8, value=eighth)
#
# for row_idx, ninth in enumerate(ninth, start=2):
#     sheet.cell(row=row_idx, column=9, value=ninth)
#
# for row_idx, tenth in enumerate(tenth, start=2):
#     sheet.cell(row=row_idx, column=10, value=tenth)
# workbook.save('output_gist.xlsx')
# workbook.close()



