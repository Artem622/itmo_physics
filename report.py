import math

import openpyxl

data = sorted(
    [715, 1050, 829, 853, 1014, 872, 886, 861, 852, 787, 857, 838, 810, 839, 808, 821, 833, 868, 847, 899, 843,
     839, 835, 826, 827, 814, 816, 819, 816, 776, 781, 773, 789, 777, 782, 795, 797, 801, 765, 796, 862, 869, 825, 796,
     809, 739, 743, 872, 811, 789, 801, 752, 784, 787, 820, 843, 892, 948, 923, 931, 895, 894, 852, 835, 861, 841, 862,
     873, 980, 888, 939, 908, 959, 874, 776, 791, 890, 855, 904, 898, 894, 901, 904, 891, 948, 888, 928, 1012,
     837, 800, 843, 864, 854, 810, 784, 835, 844, 912, 810, 844, 892, 780, 797, 827, 799, 801, 843, 819, 780, 859, 841,
     827, 806, 868, 863, 803, 779, 787, 853, 798, 813, 833, 654, 727, 823, 827, 854, 868, 846, 817, 849]
)# Ti
data = list(map(lambda point: point / 1000, data))

def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"
def average(a):
    answer = sum(a) / len(a)
    return answer

numbers = []# №
for i in range(1,len(data)+1):
    numbers.append(i)

avg =  average(data)
print(avg)
delta = []# Ti-(T)

for i in data:
    delta.append(i-avg)

delta_square = []

for i in delta:# (Ti-(T))ˆ2
    delta_square.append(i**2)
print(sum(delta))

def q(inp):
    s = 0
    for i in inp:
        s += ((i-average(inp))**2)
    answer = math.sqrt((1/(len(inp)-1))*s)
    return answer

qm = q(data)

pmax = 1/(q(data)*math.sqrt(2*math.pi))
print(pmax)
print(max(data),min(data))
print(math.sqrt(len(data)))

delta_t = toFixed(((max(data)-min(data))/(int(math.sqrt(len(data))))),4)
delta_t = float(delta_t)
print(delta_t)

first = []
second = []
third = []
forth = []
fifth = []
sixth = []
seventh = []
eighth = []
ninth = []
tenth = []
eleventh = []

min_in_inp = min(data)
max_in_inp = max(data)
for i in data:
    if i >= min_in_inp and i < min_in_inp + delta_t * 1:
        first.append(i)
    elif i >= min_in_inp + delta_t * 1 and i < min_in_inp + delta_t * 2:
        second.append(i)
    elif i >= min_in_inp +delta_t * 2 and i < min_in_inp + delta_t * 3:
        third.append(i)
    elif i >= min_in_inp +delta_t * 3 and i < min_in_inp + delta_t * 4:
        forth.append(i)
    elif i >= min_in_inp +delta_t * 4 and i < min_in_inp + delta_t * 5:
        fifth.append(i)
    elif i >= min_in_inp +delta_t * 5 and i < min_in_inp + delta_t * 6:
        sixth.append(i)
    elif i >= min_in_inp +delta_t * 6 and i < min_in_inp + delta_t * 7:
        seventh.append(i)
    elif i >= min_in_inp +delta_t * 7 and i < min_in_inp + delta_t * 8:
        eighth.append(i)
    elif i >= min_in_inp +delta_t * 8 and i < min_in_inp + delta_t * 9:
        ninth.append(i)
    elif i >= min_in_inp +delta_t * 9 and i < min_in_inp + delta_t * 10:
        tenth.append(i)
    elif i >= min_in_inp +delta_t * 10 and i <= min_in_inp + delta_t * 11:
        eleventh.append(i)

answer = first + second + third + forth + fifth +sixth + seventh + eighth +ninth + tenth
lens = [len(first), len(second), len(third), len(forth), len(fifth), len(sixth), len(seventh), len(eighth), len(ninth), len(tenth), len(eleventh)]
print(lens,'lens')
def density(inp):
    answer = len(inp)/(len(data)*delta_t)
    return answer

print(density(eleventh),'density')


def p(inp):
    return ((1/(qm*math.sqrt(2*math.pi)))*math.exp( -(((inp-avg)**2)/(2*qm**2)) ))



print(p(1.032))


print(avg+3*qm)
count = 0
for i in data:
    if (i >= 0.6645802235671479 and i <= 1.016534280249646):
        count+=1
print(count)
print(129/len(data))
print(avg,'avg')
print(qm)

pmax = 1/(q(data)*math.sqrt(2*math.pi))
print(pmax)

summ = 0
for i in data:
    summ += (i-avg)**2
qt = math.sqrt((1/(len(data)*(len(data)-1)))*summ)
print(qt,'qt')
print(qt*1.9840)
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.cell(row=1, column=1, value='№')
# sheet.cell(row=1, column=2, value='t_i , с')
# sheet.cell(row=1, column=3, value='t_i- ⟨t⟩_N, с')
# sheet.cell(row=1, column=4, value='(t_i- ⟨t⟩_N )^2, с2')
#
# for row_idx, numbers in enumerate(numbers, start=2):
#     sheet.cell(row=row_idx, column=1, value=numbers)
#
# for row_idx, data in enumerate(data, start=2):
#     sheet.cell(row=row_idx, column=2, value=data)
#
# for row_idx, delta in enumerate(delta, start=2):
#     sheet.cell(row=row_idx, column=3, value=delta)
#
# for row_idx, delta_square in enumerate(delta_square, start=2):
#     sheet.cell(row=row_idx, column=4, value=delta_square)
# workbook.save('report.xlsx')
# workbook.close()
#
# print("complete")